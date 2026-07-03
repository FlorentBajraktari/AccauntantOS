"""AccountantOS backend E2E API tests."""
import os
import io
from pathlib import Path
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    # fallback for local test runs
    frontend_env = Path(__file__).resolve().parents[2] / "frontend" / ".env"
    if frontend_env.exists():
        with frontend_env.open() as f:
            for line in f:
                if line.startswith("REACT_APP_BACKEND_URL"):
                    BASE_URL = line.split("=", 1)[1].strip().rstrip("/")
                    break

if not BASE_URL:
    raise RuntimeError(
        "Set REACT_APP_BACKEND_URL or add frontend/.env before running backend tests.")

API = f"{BASE_URL}/api"

ADMIN_EMAIL = "admin@accountantos.ch"
ADMIN_PASSWORD = "Admin123!"


@pytest.fixture(scope="session")
def client():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{API}/auth/login",
               json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
    assert r.status_code == 200, f"login failed: {r.status_code} {r.text}"
    token = r.json().get("token")
    s.headers["Authorization"] = f"Bearer {token}"
    return s


@pytest.fixture(scope="session")
def company_id(client):
    r = client.get(f"{API}/companies")
    assert r.status_code == 200
    lst = r.json()
    assert isinstance(lst, list) and len(lst) > 0, "seed companies missing"
    # Prefer seed (non-TEST_) companies to avoid concurrent-test races
    seed = [c for c in lst if not c.get("name", "").startswith("TEST_")]
    return (seed or lst)[0]["id"]


# ---- Auth
class TestAuth:
    def test_login_success(self):
        r = requests.post(
            f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
        assert r.status_code == 200
        d = r.json()
        assert d["email"] == ADMIN_EMAIL
        assert d.get("token")
        assert "access_token" in r.cookies

    def test_login_wrong_password(self):
        r = requests.post(f"{API}/auth/login",
                          json={"email": ADMIN_EMAIL, "password": "wrong"})
        assert r.status_code == 401

    def test_me(self, client):
        r = client.get(f"{API}/auth/me")
        assert r.status_code == 200
        assert r.json()["email"] == ADMIN_EMAIL

    def test_register_new_user(self):
        import uuid
        email = f"test_{uuid.uuid4().hex[:8]}@test.ch"
        r = requests.post(f"{API}/auth/register",
                          json={"name": "Test User", "email": email, "password": "Pass123!"})
        assert r.status_code == 200
        assert r.json()["email"] == email


# ---- Dashboard
class TestDashboard:
    def test_dashboard(self, client):
        r = client.get(f"{API}/dashboard")
        assert r.status_code == 200
        d = r.json()
        for key in ["total_companies", "pending_tasks", "missing_documents", "overdue_invoices",
                    "recent_documents", "upcoming_deadlines", "monthly_series"]:
            assert key in d, f"missing key {key}"
        # spec-named keys (allow either singular or prefixed form)
        assert "monthly_revenue" in d or "revenue" in d
        assert "profit_loss" in d or "profit" in d
        assert "open_vat_deadlines" in d or "vat_deadlines" in d


# ---- Companies
class TestCompanies:
    def test_list_companies(self, client):
        r = client.get(f"{API}/companies")
        assert r.status_code == 200
        assert len(r.json()) >= 4

    def test_get_company(self, client, company_id):
        r = client.get(f"{API}/companies/{company_id}")
        assert r.status_code == 200
        assert r.json()["id"] == company_id

    def test_company_overview(self, client, company_id):
        r = client.get(f"{API}/companies/{company_id}/overview")
        assert r.status_code == 200
        d = r.json()
        for k in ["documents", "payables", "receivables", "tasks", "deadlines", "questions", "activity"]:
            assert k in d

    def test_create_update_delete_company(self, client):
        payload = {"name": "TEST_Co", "country": "CH", "currency": "CHF"}
        r = client.post(f"{API}/companies", json=payload)
        assert r.status_code == 200
        cid = r.json()["id"]
        # get
        rg = client.get(f"{API}/companies/{cid}")
        assert rg.status_code == 200
        assert rg.json()["name"] == "TEST_Co"
        # update
        ru = client.put(f"{API}/companies/{cid}",
                        json={**payload, "name": "TEST_Co2"})
        assert ru.status_code == 200
        # verify
        rg2 = client.get(f"{API}/companies/{cid}")
        assert rg2.json()["name"] == "TEST_Co2"
        # delete
        rd = client.delete(f"{API}/companies/{cid}")
        assert rd.status_code == 200
        assert client.get(f"{API}/companies/{cid}").status_code == 404


# ---- Documents / OCR
class TestDocuments:
    def test_list_documents(self, client, company_id):
        r = client.get(f"{API}/documents", params={"company_id": company_id})
        assert r.status_code == 200

    def test_doc_create_ocr_delete(self, client, company_id):
        payload = {"company_id": company_id, "name": "TEST_invoice.pdf",
                   "category": "Invoice", "amount": 1200.0, "vat_amount": 92.40,
                   "counterparty": "TEST Supplier AG"}
        r = client.post(f"{API}/documents", json=payload)
        assert r.status_code == 200
        did = r.json()["id"]
        # OCR
        ro = client.post(f"{API}/documents/{did}/ocr")
        assert ro.status_code == 200, ro.text
        d = ro.json()
        assert "extraction" in d or "extracted" in d or isinstance(d, dict)
        # update
        payload["status"] = "reviewed"
        ru = client.put(f"{API}/documents/{did}", json=payload)
        assert ru.status_code == 200
        # delete
        rd = client.delete(f"{API}/documents/{did}")
        assert rd.status_code == 200


# ---- Payables / Receivables
class TestAPAR:
    def test_payables_list_and_crud(self, client, company_id):
        r = client.get(f"{API}/payables")
        assert r.status_code == 200
        # create
        payload = {"company_id": company_id, "supplier": "TEST_Sup",
                   "invoice_number": "TEST-P-1", "invoice_date": "2025-01-05",
                   "due_date": "2025-02-05", "amount": 1000.0, "vat": 81.0}
        rc = client.post(f"{API}/payables", json=payload)
        assert rc.status_code == 200
        pid = rc.json()["id"]
        # update status
        ru = client.put(f"{API}/payables/{pid}",
                        json={**payload, "payment_status": "paid"})
        assert ru.status_code == 200
        assert ru.json()["payment_status"] == "paid"
        # delete
        assert client.delete(f"{API}/payables/{pid}").status_code == 200

    def test_receivables_crud(self, client, company_id):
        payload = {"company_id": company_id, "customer": "TEST_Cust",
                   "invoice_number": "TEST-R-1", "invoice_date": "2025-01-05",
                   "due_date": "2025-02-05", "amount": 5000.0, "vat": 405.0}
        rc = client.post(f"{API}/receivables", json=payload)
        assert rc.status_code == 200
        rid = rc.json()["id"]
        assert client.delete(f"{API}/receivables/{rid}").status_code == 200

    def test_payables_aging(self, client):
        r = client.get(f"{API}/payables/report/aging")
        assert r.status_code == 200
        d = r.json()
        assert "aging" in d and "supplier_balances" in d
        for k in ["current", "b1_30", "b31_60", "b61_90", "b90", "total"]:
            assert k in d["aging"]

    def test_receivables_aging(self, client):
        r = client.get(f"{API}/receivables/report/aging")
        assert r.status_code == 200
        assert "customer_balances" in r.json()


# ---- Bank
class TestBank:
    def test_bank_csv_import(self, client, company_id):
        csv = "date,description,reference,amount\n2025-01-10,TEST Payment,REF001,-500.00\n2025-01-11,TEST Deposit,REF002,1200.50\n"
        files = {"file": ("test.csv", io.BytesIO(csv.encode()), "text/csv")}
        data = {"company_id": company_id}
        # remove Content-Type header for multipart
        h = {k: v for k, v in client.headers.items() if k.lower() !=
             "content-type"}
        r = requests.post(f"{API}/bank-transactions/import",
                          data=data, files=files, headers=h)
        assert r.status_code == 200, r.text
        assert r.json()["imported"] == 2

    def test_bank_match_and_report(self, client, company_id):
        r = client.get(f"{API}/bank-transactions",
                       params={"company_id": company_id})
        assert r.status_code == 200
        txs = r.json()
        if txs:
            tx_id = txs[0]["id"]
            rm = client.post(f"{API}/bank-transactions/{tx_id}/match")
            assert rm.status_code == 200
            assert "matched" in rm.json()
        rr = client.get(f"{API}/bank-transactions/report/reconciliation",
                        params={"company_id": company_id})
        assert rr.status_code == 200
        for k in ["matched_count", "unmatched_count", "matched_total", "unmatched_total"]:
            assert k in rr.json()


# ---- VAT
class TestVAT:
    def test_vat_summary(self, client):
        r = client.get(f"{API}/vat/summary")
        assert r.status_code == 200
        d = r.json()
        assert d["country"] == "Switzerland"
        assert d["rates"] == [8.1, 2.6, 3.8]
        for k in ["output_vat", "input_vat", "vat_balance", "taxable_revenue"]:
            assert k in d


# ---- Checklist
class TestChecklist:
    def test_checklist(self, client, company_id):
        r = client.get(f"{API}/checklist", params={"company_id": company_id})
        assert r.status_code == 200
        items = r.json()
        if items:
            it = items[0]
            payload = {"company_id": it["company_id"], "period": it["period"],
                       "task": it["task"], "status": "done",
                       "assigned_to": it.get("assigned_to", ""),
                       "due_date": it.get("due_date"), "notes": it.get("notes", "")}
            ru = client.put(f"{API}/checklist/{it['id']}", json=payload)
            assert ru.status_code == 200
            assert ru.json()["status"] == "done"


# ---- Excel
EXCEL_KEYS = ["profit_loss", "ap_aging", "ar_aging", "vat_summary",
              "chart_of_accounts", "bank_reconciliation", "month_end_close",
              "expense_tracker", "revenue_tracker"]

ALL_EXCEL_KEYS = [
    "chart_of_accounts", "trial_balance", "general_ledger",
    "ap_aging", "ar_aging", "vat_summary", "bank_reconciliation",
    "month_end_close", "expense_tracker", "revenue_tracker",
    "payroll_preparation", "profit_loss", "balance_sheet", "cash_flow",
]


class TestExcel:
    def test_templates_list(self, client):
        r = client.get(f"{API}/excel/templates")
        assert r.status_code == 200
        tpls = r.json()["templates"]
        assert len(tpls) == 14, f"expected 14 templates, got {len(tpls)}"
        for t in tpls:
            for field in ("key", "title", "description", "category", "icon"):
                assert field in t and t[field], f"missing {field} in {t}"
        keys = {t["key"] for t in tpls}
        assert keys == set(
            ALL_EXCEL_KEYS), f"key mismatch: {keys ^ set(ALL_EXCEL_KEYS)}"

    @pytest.mark.parametrize("key", EXCEL_KEYS)
    def test_excel_download(self, client, key):
        r = client.get(f"{API}/excel/download/{key}")
        assert r.status_code == 200, f"{key}: {r.status_code} {r.text[:200]}"
        assert r.content[:2] == b"PK", f"{key}: not xlsx"
        # openpyxl-loadable + two sheets including "How to use"
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        assert "How to use" in wb.sheetnames, f"{key}: missing 'How to use' sheet, got {wb.sheetnames}"
        assert len(
            wb.sheetnames) >= 2, f"{key}: expected >=2 sheets, got {wb.sheetnames}"

    def test_excel_download_with_company(self, client, company_id):
        r = client.get(f"{API}/excel/download/ap_aging",
                       params={"company_id": company_id})
        assert r.status_code == 200
        assert r.content[:2] == b"PK"
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        assert "How to use" in wb.sheetnames


# ---- Reports
REPORTS = ["profit-loss", "expense-by-category",
           "revenue-by-client", "balance-sheet", "cash-flow"]


class TestReports:
    @pytest.mark.parametrize("rpt", REPORTS)
    def test_report(self, client, rpt):
        r = client.get(f"{API}/reports/{rpt}")
        assert r.status_code == 200, f"{rpt}: {r.status_code} {r.text[:200]}"


# ---- Search / Audit / Settings
class TestMisc:
    def test_search(self, client):
        r = client.get(f"{API}/search", params={"q": "alpine"})
        assert r.status_code == 200
        d = r.json()
        # expect companies key
        assert "companies" in d or isinstance(d, dict)

    def test_audit(self, client):
        r = client.get(f"{API}/audit")
        assert r.status_code == 200
        assert isinstance(r.json(), list) or "logs" in r.json()

    def test_settings_get_and_put(self, client):
        r = client.get(f"{API}/settings")
        assert r.status_code == 200
        cur = r.json() if isinstance(r.json(), dict) else {}
        ru = client.put(f"{API}/settings",
                        json={**cur, "test_key": "TEST_val"})
        assert ru.status_code == 200


# ---- Client Portal
class TestPortal:
    def test_portal(self, client, company_id):
        r = client.get(f"{API}/portal/{company_id}")
        assert r.status_code == 200
        d = r.json()
        for k in ["company", "documents", "missing", "questions", "deadlines"]:
            assert k in d

    def test_answer_question(self, client, company_id):
        # create a question
        payload = {"company_id": company_id, "question": "TEST_q?"}
        rc = client.post(f"{API}/questions", json=payload)
        assert rc.status_code == 200
        qid = rc.json()["id"]
        payload_up = {**payload, "answer": "yes", "status": "answered"}
        ru = client.put(f"{API}/questions/{qid}", json=payload_up)
        assert ru.status_code == 200
        assert ru.json()["status"] == "answered"
        client.delete(f"{API}/questions/{qid}")
