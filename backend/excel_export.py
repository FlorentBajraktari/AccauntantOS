"""Professional, modern Excel template generation using openpyxl.

Modern design:
- Branded header banner (blue fill, white title) + subtitle + meta line.
- Native Excel Tables (banded rows, built-in filter dropdowns, auto totals row).
- Currency / date / percent number formats, right-aligned numerics.
- Data-validation dropdowns for status columns.
- Frozen header, colored sheet tab, print-ready page setup (fit to width,
  landscape for wide sheets, header/footer with page numbers).
- Redesigned "How to use" instructions sheet.
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.utils import get_column_letter

BRAND = "0055FF"
INK = "0F172A"
SLATE = "475569"
MUTED = "94A3B8"
WHITE = "FFFFFF"

FONT = "Calibri"
CURRENCY_FMT = '#,##0.00 "CHF";[Red]-#,##0.00 "CHF"'
DATE_FMT = "DD.MM.YYYY"
PCT_FMT = '0.0"%"'


def _banner(ws, title, subtitle, ncols):
    last = get_column_letter(max(ncols, 1))
    # Brand banner (title + subtitle) rows 1-2
    ws.merge_cells(f"A1:{last}1")
    ws.merge_cells(f"A2:{last}2")
    for r in (1, 2):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BRAND)
    tc = ws["A1"]
    tc.value = title
    tc.font = Font(name=FONT, bold=True, size=20, color=WHITE)
    tc.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 32
    sc = ws["A2"]
    sc.value = subtitle
    sc.font = Font(name=FONT, size=10.5, italic=True, color="DBEAFE")
    sc.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 20
    # Meta line (row 3)
    ws.merge_cells(f"A3:{last}3")
    mc = ws["A3"]
    mc.value = f"AccountantOS   ·   Generated {datetime.now().strftime('%d.%m.%Y %H:%M')}   ·   Currency CHF   ·   Switzerland"
    mc.font = Font(name=FONT, size=9, color=MUTED)
    mc.alignment = Alignment(vertical="center", horizontal="right", indent=1)
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 6  # spacer


def _instructions_sheet(wb, name, lines):
    ws = wb.create_sheet("How to use")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "94A3B8"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 98
    ws.merge_cells("B1:C1")
    b = ws["B1"]
    b.value = f"  {name} — How to use this template"
    b.fill = PatternFill("solid", fgColor=BRAND)
    b.font = Font(name=FONT, bold=True, size=15, color=WHITE)
    b.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells("B2:C2")
    ws["B2"].value = "AccountantOS · Professional Swiss Accounting Platform"
    ws["B2"].font = Font(name=FONT, size=10, italic=True, color=SLATE)
    r = 4
    for i, line in enumerate(lines, start=1):
        n = ws.cell(row=r, column=2, value=str(i))
        n.font = Font(name=FONT, bold=True, size=11, color=WHITE)
        n.fill = PatternFill("solid", fgColor=BRAND)
        n.alignment = Alignment(horizontal="center", vertical="center")
        tcell = ws.cell(row=r, column=3, value=line)
        tcell.font = Font(name=FONT, size=11, color=INK)
        tcell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 26
        r += 1
    tip = ws.cell(row=r + 1, column=3, value="Tip: rows are a native Excel table — use the header filters, and the TOTAL row updates automatically.")
    tip.font = Font(name=FONT, size=9.5, italic=True, color=SLATE)


def _page_setup(ws, ncols, title):
    ws.page_setup.orientation = "landscape" if ncols > 6 else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.6
    ws.oddHeader.left.text = "AccountantOS"
    ws.oddHeader.center.text = title
    ws.oddFooter.left.text = "&D"
    ws.oddFooter.right.text = "Page &P of &N"


TEMPLATES = {
    "chart_of_accounts": {
        "title": "Chart of Accounts", "sub": "Swiss KMU standard account structure", "category": "Ledgers & Accounts", "icon": "list-tree",
        "headers": ["Account No.", "Account Name", "Type", "Category", "VAT Code", "Opening Balance", "Notes"],
        "fmt": {6: CURRENCY_FMT}, "totals": [], "dropdowns": {3: ["Asset", "Liability", "Equity", "Revenue", "Expense"]},
        "instructions": ["Fill in your account numbers following the Swiss KMU chart.", "Use the Type dropdown to classify each account.", "Opening Balance is in CHF.", "Do not delete the header row."],
    },
    "trial_balance": {
        "title": "Trial Balance", "sub": "Debit / Credit balances by account", "category": "Ledgers & Accounts", "icon": "scale",
        "headers": ["Account No.", "Account Name", "Debit", "Credit", "Balance"],
        "fmt": {3: CURRENCY_FMT, 4: CURRENCY_FMT, 5: CURRENCY_FMT}, "totals": [3, 4, 5], "dropdowns": {},
        "instructions": ["Enter debit and credit balances per account.", "Total debits must equal total credits.", "Balance column = Debit - Credit."],
    },
    "general_ledger": {
        "title": "General Ledger", "sub": "Chronological posting journal", "category": "Ledgers & Accounts", "icon": "book-open",
        "headers": ["Date", "Account No.", "Description", "Reference", "Debit", "Credit", "Running Balance"],
        "fmt": {1: DATE_FMT, 5: CURRENCY_FMT, 6: CURRENCY_FMT, 7: CURRENCY_FMT}, "totals": [5, 6], "dropdowns": {},
        "instructions": ["Post each transaction on a new row in date order.", "Every entry must have a matching debit and credit.", "Use the Reference column for document numbers."],
    },
    "ap_aging": {
        "title": "Accounts Payable Aging", "sub": "Outstanding supplier invoices by age bucket", "category": "Payables & Receivables", "icon": "arrow-down-circle",
        "headers": ["Supplier", "Invoice No.", "Invoice Date", "Due Date", "Total", "Current", "1-30", "31-60", "61-90", "90+"],
        "fmt": {3: DATE_FMT, 4: DATE_FMT, 5: CURRENCY_FMT, 6: CURRENCY_FMT, 7: CURRENCY_FMT, 8: CURRENCY_FMT, 9: CURRENCY_FMT, 10: CURRENCY_FMT},
        "totals": [5, 6, 7, 8, 9, 10], "dropdowns": {},
        "instructions": ["Aging buckets are days past due.", "Totals row summarises exposure per bucket.", "Review the 90+ column urgently."],
    },
    "ar_aging": {
        "title": "Accounts Receivable Aging", "sub": "Outstanding customer invoices by age bucket", "category": "Payables & Receivables", "icon": "arrow-up-circle",
        "headers": ["Customer", "Invoice No.", "Invoice Date", "Due Date", "Total", "Current", "1-30", "31-60", "61-90", "90+"],
        "fmt": {3: DATE_FMT, 4: DATE_FMT, 5: CURRENCY_FMT, 6: CURRENCY_FMT, 7: CURRENCY_FMT, 8: CURRENCY_FMT, 9: CURRENCY_FMT, 10: CURRENCY_FMT},
        "totals": [5, 6, 7, 8, 9, 10], "dropdowns": {},
        "instructions": ["Aging buckets are days past due.", "Follow up on overdue balances.", "Send reminders for 30+ day items."],
    },
    "vat_summary": {
        "title": "VAT Summary", "sub": "Swiss VAT reconciliation (rates 8.1% / 2.6% / 3.8%)", "category": "VAT & Bank", "icon": "receipt",
        "headers": ["Period", "Taxable Revenue", "Output VAT", "Deductible Expenses", "Input VAT", "VAT Balance", "Status"],
        "fmt": {2: CURRENCY_FMT, 3: CURRENCY_FMT, 4: CURRENCY_FMT, 5: CURRENCY_FMT, 6: CURRENCY_FMT},
        "totals": [2, 3, 4, 5, 6], "dropdowns": {7: ["Draft", "Filed", "Paid"]},
        "instructions": ["Output VAT = VAT collected on sales.", "Input VAT = VAT paid on purchases.", "VAT Balance = Output VAT - Input VAT (payable if positive)."],
    },
    "bank_reconciliation": {
        "title": "Bank Reconciliation", "sub": "Match bank statement lines to booked entries", "category": "VAT & Bank", "icon": "landmark",
        "headers": ["Date", "Description", "Reference", "Bank Amount", "Book Amount", "Difference", "Status"],
        "fmt": {1: DATE_FMT, 4: CURRENCY_FMT, 5: CURRENCY_FMT, 6: CURRENCY_FMT}, "totals": [4, 5, 6],
        "dropdowns": {7: ["Matched", "Unmatched", "Review"]},
        "instructions": ["Import your bank statement lines here.", "Difference should be 0 when fully reconciled.", "Mark each line Matched or Unmatched."],
    },
    "month_end_close": {
        "title": "Month-End Close Checklist", "sub": "Standard close procedure", "category": "Operations", "icon": "check-square",
        "headers": ["Task", "Assigned To", "Due Date", "Status", "Notes"],
        "fmt": {3: DATE_FMT}, "totals": [], "dropdowns": {4: ["Pending", "In Progress", "Completed", "Blocked"]},
        "instructions": ["Complete each task before closing the period.", "Assign an owner and due date to every item.", "Do not close until all items are Completed."],
    },
    "expense_tracker": {
        "title": "Expense Tracker", "sub": "Categorised business expenses", "category": "Operations", "icon": "trending-down",
        "headers": ["Date", "Supplier", "Category", "Description", "Net", "VAT", "Gross", "Payment Method"],
        "fmt": {1: DATE_FMT, 5: CURRENCY_FMT, 6: CURRENCY_FMT, 7: CURRENCY_FMT}, "totals": [5, 6, 7],
        "dropdowns": {8: ["Bank Transfer", "Credit Card", "Cash", "Direct Debit"]},
        "instructions": ["Log every business expense.", "Gross = Net + VAT.", "Keep receipts for all entries."],
    },
    "revenue_tracker": {
        "title": "Revenue Tracker", "sub": "Sales income by client", "category": "Operations", "icon": "trending-up",
        "headers": ["Date", "Customer", "Invoice No.", "Description", "Net", "VAT", "Gross", "Status"],
        "fmt": {1: DATE_FMT, 5: CURRENCY_FMT, 6: CURRENCY_FMT, 7: CURRENCY_FMT}, "totals": [5, 6, 7],
        "dropdowns": {8: ["Draft", "Sent", "Paid", "Overdue"]},
        "instructions": ["Record all revenue transactions.", "Gross = Net + VAT.", "Update Status as invoices are paid."],
    },
    "payroll_preparation": {
        "title": "Payroll Preparation", "sub": "Monthly employee payroll", "category": "Operations", "icon": "users",
        "headers": ["Employee", "Gross Salary", "AHV/IV/EO", "ALV", "Pension (BVG)", "Net Salary", "Notes"],
        "fmt": {2: CURRENCY_FMT, 3: CURRENCY_FMT, 4: CURRENCY_FMT, 5: CURRENCY_FMT, 6: CURRENCY_FMT},
        "totals": [2, 3, 4, 5, 6], "dropdowns": {},
        "instructions": ["Enter gross salary and social deductions.", "Net Salary = Gross - deductions.", "Swiss social contributions: AHV/IV/EO, ALV, BVG."],
    },
    "profit_loss": {
        "title": "Profit & Loss Statement", "sub": "Income statement", "category": "Financial Statements", "icon": "bar-chart-3",
        "headers": ["Line Item", "Category", "Current Period", "Prior Period", "Variance"],
        "fmt": {3: CURRENCY_FMT, 4: CURRENCY_FMT, 5: CURRENCY_FMT}, "totals": [3, 4, 5],
        "dropdowns": {2: ["Revenue", "COGS", "Operating Expense", "Other"]},
        "instructions": ["List revenue then expenses.", "Net Profit = Revenue - Expenses.", "Variance = Current - Prior."],
    },
    "balance_sheet": {
        "title": "Balance Sheet", "sub": "Statement of financial position", "category": "Financial Statements", "icon": "scale",
        "headers": ["Line Item", "Section", "Current Period", "Prior Period"],
        "fmt": {3: CURRENCY_FMT, 4: CURRENCY_FMT}, "totals": [3, 4],
        "dropdowns": {2: ["Assets", "Liabilities", "Equity"]},
        "instructions": ["Assets = Liabilities + Equity.", "Group items by section.", "Ensure the balance sheet balances."],
    },
    "cash_flow": {
        "title": "Cash Flow Statement", "sub": "Sources and uses of cash", "category": "Financial Statements", "icon": "waves",
        "headers": ["Line Item", "Activity", "Amount"],
        "fmt": {3: CURRENCY_FMT}, "totals": [3],
        "dropdowns": {2: ["Operating", "Investing", "Financing"]},
        "instructions": ["Classify each flow by activity.", "Net Cash Flow = sum of all activities.", "Reconcile to opening/closing cash."],
    },
}


def build_template(key: str, rows: list | None = None) -> bytes:
    spec = TEMPLATES[key]
    headers = spec["headers"]
    ncols = len(headers)
    fmt = spec["fmt"]
    totals = spec["totals"]

    wb = Workbook()
    ws = wb.active
    ws.title = spec["title"][:31]
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BRAND

    _banner(ws, spec["title"], spec["sub"], ncols)

    header_row = 5
    for c, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=title)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 26

    # Column widths
    for c, title in enumerate(headers, start=1):
        w = 16
        if c in fmt and fmt[c] == CURRENCY_FMT:
            w = 16
        if "Name" in title or "Description" in title or "Notes" in title or "Task" in title or "Item" in title:
            w = 26
        ws.column_dimensions[get_column_letter(c)].width = w

    rows = rows or []
    data_start = header_row + 1
    n_data = max(len(rows), 20)
    data_end = data_start + n_data - 1

    for r_off in range(n_data):
        r_idx = data_start + r_off
        row = rows[r_off] if r_off < len(rows) else None
        for c_idx, title in enumerate(headers, start=1):
            val = row.get(title) if isinstance(row, dict) else None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name=FONT, size=10, color=INK)
            if c_idx in fmt:
                cell.number_format = fmt[c_idx]
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[r_idx].height = 18

    # Totals row (Excel Table totals)
    table_end = data_end
    if totals:
        total_row = data_end + 1
        table_end = total_row
        for c_idx in totals:
            cell = ws.cell(row=total_row, column=c_idx)
            cell.number_format = fmt.get(c_idx, CURRENCY_FMT)

    # Native Excel Table (banded rows + filter + optional totals row)
    ref = f"A{header_row}:{get_column_letter(ncols)}{table_end}"
    table = Table(displayName=f"tbl_{key}", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    cols = []
    for i, h in enumerate(headers, start=1):
        tcol = TableColumn(id=i, name=h)
        if totals:
            if i == 1:
                tcol.totalsRowLabel = "TOTAL"
            elif i in totals:
                tcol.totalsRowFunction = "sum"
        cols.append(tcol)
    table.tableColumns = cols
    if totals:
        table.totalsRowShown = True
        table.totalsRowCount = 1
    ws.add_table(table)

    # Data validation dropdowns
    for col_idx, options in spec.get("dropdowns", {}).items():
        dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        col = get_column_letter(col_idx)
        dv.add(f"{col}{data_start}:{col}{data_end}")

    ws.freeze_panes = ws.cell(row=data_start, column=1)
    _page_setup(ws, ncols, spec["title"])

    _instructions_sheet(wb, spec["title"], spec["instructions"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def list_templates():
    return [
        {"key": k, "title": v["title"], "description": v["sub"], "category": v["category"], "icon": v["icon"]}
        for k, v in TEMPLATES.items()
    ]
